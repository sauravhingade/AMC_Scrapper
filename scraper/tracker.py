"""
Writes a single Excel file tracking each AMC's scrape status for the
most recent run: pass/fail, the paths of any files that were
downloaded, and the error message if it failed.

Called once, after every AMC in the batch has been attempted. The file
is fully overwritten on each call -- it's a snapshot of the latest run,
not an appended history.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TRACKING_XLSX_PATH = "amc_tracking_status.xlsx"

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(name="Arial", color="006100")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAIL_FONT = Font(name="Arial", color="9C0006")
BODY_FONT = Font(name="Arial")

HEADERS = ["AMC Name", "URL", "Status", "Downloaded Files", "Error", "Last Run"]

COLUMN_WIDTHS = {1: 28, 2: 45, 3: 10, 4: 55, 5: 45, 6: 20}


def write_tracking_excel(results, path=TRACKING_XLSX_PATH):
    """
    results: list of dicts, one per AMC, shaped like:
        {
            "name": "Axis Mutual Fund",
            "url": "https://...",
            "status": "PASS" or "FAIL",
            "files": ["downloads/Axis_....pdf", ...],
            "error": "" (or the exception message on failure),
        }

    Overwrites `path` completely every call.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AMC Tracking"

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    run_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # noqa: DTZ005

    for row_idx, result in enumerate(results, start=2):
        files_str = "\n".join(result.get("files") or [])
        status = result.get("status", "FAIL")

        ws.append(
            [
                result.get("name", ""),
                result.get("url", ""),
                status,
                files_str,
                result.get("error", ""),
                run_timestamp,
            ]
        )

        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 4))

        status_cell = ws.cell(row=row_idx, column=3)
        if status == "PASS":
            status_cell.fill = PASS_FILL
            status_cell.font = PASS_FONT
        else:
            status_cell.fill = FAIL_FILL
            status_cell.font = FAIL_FONT

    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"\nTracking sheet written to: {os.path.abspath(path)}")
